from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Avg, Sum, Count, Q
from django import forms
from django.urls import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from urllib.parse import quote
from django.utils import timezone
import openpyxl

from.models import AcademicYear, Term, ClassRoom, Subject, Teacher, Exam, ExamResult, StudentReport
from students.models import Student
from.forms import AcademicYearForm, TermForm, SubjectForm, ExamForm, ClassRoomForm
from accounts.views import is_teacher

# ========== FORMS ==========
class TeacherForm(forms.ModelForm):
    class Meta:
        model = Teacher
        fields = ['first_name', 'last_name', 'email', 'phone', 'photo']
        widgets = {
            'first_name': forms.TextInput(attrs={'class': 'form-control'}),
            'last_name': forms.TextInput(attrs={'class': 'form-control'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
            'phone': forms.TextInput(attrs={'class': 'form-control'}),
            'photo': forms.FileInput(attrs={'class': 'form-control'}),
        }

# ========== NECTA GRADING SYSTEM ==========
def get_pre_primary_grade(score):
    if score is None: return 'F'
    score = float(score)
    if score >= 75: return 'A'
    elif score >= 65: return 'B'
    elif score >= 45: return 'C'
    elif score >= 30: return 'D'
    else: return 'F'

def get_ordinary_points(score):
    score = float(score or 0)
    if score >= 75: return 1
    elif score >= 65: return 2
    elif score >= 45: return 3
    elif score >= 30: return 4
    else: return 9

def get_advanced_points(score):
    score = float(score or 0)
    if score >= 75: return 1
    elif score >= 65: return 2
    elif score >= 45: return 3
    elif score >= 30: return 4
    else: return 5

def calculate_best_points(all_points, level_group):
    """Returns: best_sum, division_name. Forces Best 7 for O, Best 3 for A"""
    if not all_points: return 0, '-'
    if level_group == 'ORDINARY':
        if 9 in all_points:
            return sum(all_points), 'Division 0'
        best = sorted(all_points)[:7]
        total = sum(best)
        if 7 <= total <= 17: div = 'Division I'
        elif 18 <= total <= 21: div = 'Division II'
        elif 22 <= total <= 25: div = 'Division III'
        elif 26 <= total <= 34: div = 'Division IV'
        else: div = 'Division 0'
        return total, div

    if level_group == 'ADVANCED':
        points_no_gs = [p for p in all_points if p!= 5]
        if 5 in all_points:
            return sum(all_points), 'Division 0'
        best = sorted(points_no_gs)[:3]
        total = sum(best)
        if 3 <= total <= 9: div = 'Division I'
        elif 10 <= total <= 11: div = 'Division II'
        elif 12 <= total <= 13: div = 'Division III'
        elif 14 <= total <= 17: div = 'Division IV'
        else: div = 'Division 0'
        return total, div
    return 0, '-'
# ========== END GRADING ==========

def _get_student_name(student):
    return f"{student.first_name} {student.middle_name or ''} {student.last_name}".strip()

# ========== DASHBOARD ==========
@login_required
@user_passes_test(is_teacher)
def class_list(request):
    context = {
        'classes_count': ClassRoom.objects.count(),
        'students_count': Student.objects.filter(status="ACTIVE", is_active=True).count(),
        'teachers_count': Teacher.objects.count(),
        'subjects_count': Subject.objects.count(),
        'exams_count': Exam.objects.count(),
        'active_year': AcademicYear.objects.filter(is_active=True).first(),
        'classes': ClassRoom.objects.select_related('class_teacher').annotate(
            student_count=Count('students', filter=Q(students__status="ACTIVE", students__is_active=True))
        ).order_by('level', 'stream')
    }
    return render(request, "academics/class_list.html", context)

# ========== CLASS VIEWS ==========
def class_create(request):
    form = ClassRoomForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Class created successfully')
        return redirect('academics:class_list')
    return render(request, "academics/class_form.html", {'form': form, 'title': 'Add Class'})

def class_dashboard(request, pk):
    cls = get_object_or_404(ClassRoom.objects.select_related('class_teacher'), pk=pk)
    students = cls.students.filter(status="ACTIVE", is_active=True).order_by('first_name')
    subjects = Subject.objects.filter(class_room=cls).select_related('teacher')
    academic_year = AcademicYear.objects.filter(is_active=True).first()

    context = {
        'cls': cls,
        'students': students,
        'subjects': subjects,
        'student_count': students.count(),
        'subject_count': subjects.count(),
        'academic_year': academic_year,
    }
    return render(request, 'academics/class_dashboard.html', context)

def class_overview(request, class_id):
    class_room = get_object_or_404(ClassRoom, pk=class_id)
    students = Student.objects.filter(class_room=class_room, status="ACTIVE", is_active=True).order_by('first_name', 'last_name')
    subjects = Subject.objects.filter(class_room=class_room)
    academic_year = AcademicYear.objects.filter(is_active=True).first()
    teachers = Teacher.objects.filter(subjects__in=subjects).distinct()
    exams = Exam.objects.filter(class_room=class_room, term__academic_year=academic_year).select_related('term').order_by('term__id', 'id')
    latest_exam = exams.last()

    term_mapping = {1: "TERM 1", 2: "TERM 2", 3: "TERM 3", 4: "TERM 4"}
    term_exams = {}
    for term_num, term_label in term_mapping.items():
        term_exams[term_label] = []
        for exam in [e for e in exams if e.term.id == term_num]:
            reports = StudentReport.objects.filter(exam=exam).select_related('student').order_by('-average_marks')
            for pos, r in enumerate(reports, 1):
                r.position = pos
                r.total_students = students.count()
            term_exams[term_label].append({'exam': exam, 'reports': reports})

    context = {
        'class_room': class_room,
        'students': students,
        'teachers': teachers,
        'subjects': subjects,
        'student_count': students.count(),
        'academic_year': academic_year,
        'term_exams': term_exams,
        'latest_exam': latest_exam
    }
    return render(request, "academics/class_overview.html", context)

def class_detail(request, pk):
    class_room = get_object_or_404(ClassRoom, id=pk)
    students = Student.objects.filter(class_room=class_room).order_by('first_name')
    exams = Exam.objects.filter(class_room=class_room, is_active=True)
    context = {'class_room': class_room, 'students': students, 'exams': exams}
    return render(request, 'academics/class_detail.html', context)

# ========== YEAR VIEWS ==========
def year_list(request):
    years = AcademicYear.objects.all()
    return render(request, "academics/year_list.html", {'years': years})

def year_create(request):
    if request.method == "POST":
        form = AcademicYearForm(request.POST)
        if form.is_valid():
            if form.cleaned_data['is_active']:
                AcademicYear.objects.update(is_active=False)
            year = form.save()
            messages.success(request, f"Academic Year {year.name} created!")
            return redirect('academics:year_list')
    else:
        form = AcademicYearForm()
    return render(request, "academics/year_form.html", {'form': form, 'title': 'Add Academic Year'})

def year_update(request, pk):
    year = get_object_or_404(AcademicYear, pk=pk)
    if request.method == "POST":
        form = AcademicYearForm(request.POST, instance=year)
        if form.is_valid():
            if form.cleaned_data['is_active']:
                AcademicYear.objects.exclude(pk=pk).update(is_active=False)
            form.save()
            messages.success(request, f"Academic Year {year.name} updated!")
            return redirect('academics:year_list')
    else:
        form = AcademicYearForm(instance=year)
    return render(request, "academics/year_form.html", {'form': form, 'title': 'Edit Academic Year'})

def year_activate(request, pk):
    AcademicYear.objects.update(is_active=False)
    year = get_object_or_404(AcademicYear, pk=pk)
    year.is_active = True
    year.save()
    messages.success(request, f"{year.name} is now Active")
    return redirect('academics:year_list')

# ========== TERM VIEWS ==========
def term_list(request):
    terms = Term.objects.select_related('academic_year').all().order_by('-start_date')
    today = timezone.now().date()
    for t in terms:
        t.total_days = (t.end_date - t.start_date).days + 1
        t.days_passed = (today - t.start_date).days + 1
        if t.days_passed < 0: t.days_passed = 0
        if t.days_passed > t.total_days: t.days_passed = t.total_days
    return render(request, 'academics/term_list.html', {'terms': terms})

def term_create(request):
    form = TermForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Term created successfully')
        return redirect('academics:term_list')
    return render(request, "academics/term_form.html", {'form': form, 'title': 'Add Term'})

def term_update(request, pk):
    term = get_object_or_404(Term, pk=pk)
    form = TermForm(request.POST or None, instance=term)
    if form.is_valid():
        form.save()
        messages.success(request, 'Term updated successfully')
        return redirect('academics:term_list')
    return render(request, "academics/term_form.html", {'form': form, 'title': 'Edit Term'})

def term_delete(request, pk):
    term = get_object_or_404(Term, pk=pk)
    if request.method == 'POST':
        term.delete()
        messages.success(request, 'Term deleted successfully')
        return redirect('academics:term_list')
    return render(request, "academics/term_confirm_delete.html", {'term': term})

# ========== SUBJECT VIEWS ==========
def subject_list(request):
    subjects = Subject.objects.select_related('class_room', 'teacher').all()
    context = {
        'subjects': subjects,
        'assigned_count': subjects.filter(teacher__isnull=False).count(),
        'unassigned_count': subjects.filter(teacher__isnull=True).count(),
    }
    return render(request, 'academics/subject_list.html', context)

def subject_create(request):
    if request.method == "POST":
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Subject created!")
            return redirect('academics:subject_list')
    else:
        form = SubjectForm()
    return render(request, "academics/subject_form.html", {'form': form, 'title': 'Add Subject'})

def subject_update(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    form = SubjectForm(request.POST or None, instance=subject)
    if form.is_valid():
        form.save()
        messages.success(request, 'Subject updated successfully')
        return redirect('academics:subject_list')
    return render(request, "academics/subject_form.html", {'form': form, 'title': 'Edit Subject'})

def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete()
        messages.success(request, 'Subject deleted successfully')
        return redirect('academics:subject_list')
    return render(request, "academics/subject_confirm_delete.html", {'subject': subject})

# ========== TEACHER VIEWS ==========
def teacher_list(request):
    teachers = Teacher.objects.all().annotate(
        subject_count=Count('subjects', distinct=True),
        classroom_count=Count('classrooms', distinct=True)
    )
    context = {
        'teachers': teachers,
        'subjects_count': Subject.objects.count(),
        'classes_count': ClassRoom.objects.count(),
    }
    return render(request, 'academics/teacher_list.html', context)

def teacher_detail(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    return render(request, 'academics/teacher_detail.html', {'teacher': teacher})

def teacher_create(request):
    if request.method == "POST":
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Teacher created!")
            return redirect('academics:teacher_list')
    else:
        form = TeacherForm()
    return render(request, "academics/teacher_form.html", {'form': form, 'title': 'Add Teacher'})

def teacher_update(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES, instance=teacher)
        if form.is_valid():
            form.save()
            messages.success(request, f'{teacher.first_name} {teacher.last_name} updated successfully!')
            return redirect('academics:teacher_detail', pk=teacher.pk)
    else:
        form = TeacherForm(instance=teacher)
    return render(request, 'academics/teacher_form.html', {'form': form, 'teacher': teacher, 'title': 'Edit Teacher'})

def teacher_dashboard(request):
    teacher = Teacher.objects.filter(email=request.user.email).first()
    if not teacher:
        return render(request, 'academics/no_teacher_profile.html')
    my_classes = ClassRoom.objects.filter(class_teacher=teacher)
    my_subjects = Subject.objects.filter(teacher=teacher).select_related('class_room')
    active_exams = Exam.objects.filter(is_active=True)
    context = {
        'teacher': teacher,
        'my_classes': my_classes,
        'my_subjects': my_subjects,
        'active_exams': active_exams,
    }
    return render(request, 'academics/teacher_dashboard.html', context)

def assign_subject_to_teacher(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        subject = get_object_or_404(Subject, id=subject_id)
        subject.teacher = teacher
        subject.save()
        messages.success(request, f'{subject.name} assigned to {teacher.first_name}')
        return redirect('academics:teacher_detail', pk=teacher.pk)
    available_subjects = Subject.objects.filter(teacher__isnull=True).select_related('class_room')
    return render(request, 'academics/assign_subject.html', {'teacher': teacher, 'subjects': available_subjects})

# ========== EXAM VIEWS ==========
def exam_list(request):
    exams = Exam.objects.select_related('term', 'class_room').all()
    return render(request, "academics/exam_list.html", {'exams': exams})

def exam_create(request):
    if request.method == "POST":
        form = ExamForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam created!")
            return redirect('academics:exam_list')
    else:
        form = ExamForm()
    return render(request, "academics/exam_form.html", {'form': form, 'title': 'Add Exam'})

def download_template(request, exam_id):
    exam = get_object_or_404(Exam, pk=exam_id)
    subjects = Subject.objects.filter(class_room=exam.class_room).order_by('name')
    students = Student.objects.filter(class_room=exam.class_room, status="ACTIVE", is_active=True).order_by('first_name', 'last_name')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = ['AdmissionNumber', 'StudentName']
    for subject in subjects:
        headers.append(f"{subject.name}_Test")
        headers.append(f"{subject.name}_Exam")
    ws.append(headers)
    for student in students:
        row = [student.admission_number, _get_student_name(student)]
        [row.extend(['', '']) for _ in subjects]
        ws.append(row)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{exam.name}_{exam.class_room}_Template.xlsx".replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def bulk_upload_results(request, exam_id):
    exam = get_object_or_404(Exam, pk=exam_id)
    subjects = list(Subject.objects.filter(class_room=exam.class_room).order_by('name'))
    if not subjects:
        messages.error(request, "No subjects found for this class.")
        return redirect('academics:class_dashboard', pk=exam.class_room.id)
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Please select a file")
            return redirect('academics:bulk_upload_results', exam_id=exam_id)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            adm_no = row[0]
            if not adm_no: continue
            student = Student.objects.filter(admission_number=adm_no, class_room=exam.class_room, status="ACTIVE", is_active=True).first()
            if not student: continue
            col_index = 2
            ExamResult.objects.filter(student=student, exam=exam).delete()
            for subject in subjects:
                test_marks = float(row[col_index]) if row[col_index] not in [None, ''] else 0
                exam_marks = float(row[col_index + 1]) if row[col_index + 1] not in [None, ''] else 0
                col_index += 2
                ExamResult.objects.create(student=student, exam=exam, subject=subject, test_marks=test_marks, exam_marks=exam_marks)
            results = ExamResult.objects.filter(student=student, exam=exam)
            all_points = []
            points_for_best = []
            for result in results:
                avg = float(result.average_marks)
                if exam.class_room.level_group == 'ORDINARY':
                    p = get_ordinary_points(avg)
                    result.points = p
                    all_points.append(p)
                    points_for_best.append(p)
                elif exam.class_room.level_group == 'ADVANCED':
                    p = get_advanced_points(avg)
                    result.points = p
                    all_points.append(p)
                    if result.subject.name.upper()!= 'GENERAL STUDIES':
                        points_for_best.append(p)
                result.save()
            best_sum, division = calculate_best_points(points_for_best, exam.class_room.level_group)
            total = results.aggregate(Sum('total_marks'))['total_marks__sum'] or 0
            avg = results.aggregate(Avg('average_marks'))['average_marks__avg'] or 0
            total_all = sum(all_points)
            StudentReport.objects.filter(student=student, exam=exam).delete()
            StudentReport.objects.create(
                student=student, exam=exam, academic_year=exam.term.academic_year,
                total_marks=total, average_marks=avg,
                total_points=total_all,
                division_points=best_sum,
                division=division
            )
        reports = StudentReport.objects.filter(exam=exam).order_by('-average_marks')
        for position, report in enumerate(reports, start=1):
            report.position = position
            report.save()
        messages.success(request, f"Results uploaded successfully for {exam.name}")
        return redirect('academics:class_dashboard', pk=exam.class_room.id)
    return render(request, "academics/bulk_upload.html", {'exam': exam})

# ========== REPORT VIEWS ==========
def student_profile(request, pk):
    student = get_object_or_404(Student, pk=pk)
    reports = StudentReport.objects.filter(student=student).select_related('exam', 'exam__term', 'academic_year').order_by('-exam__term__academic_year__name', '-exam__term__id')
    return render(request, "students/student_profile.html", {'student': student, 'reports': reports})

def report_detail(request, pk):
    report = get_object_or_404(StudentReport, pk=pk)
    student = report.student
    class_room = student.class_room
    academic_year = report.academic_year
    total_students = Student.objects.filter(class_room=student.class_room, status="ACTIVE", is_active=True).count()
    exam_reports = StudentReport.objects.filter(exam=report.exam).order_by('-average_marks')
    for pos, r in enumerate(exam_reports, 1):
        if r.id == report.id:
            report.position = pos
            break
    report.total_students = total_students
    progress_data = StudentReport.objects.filter(student=student, academic_year=academic_year).select_related('exam', 'exam__term').order_by('exam__term__id')
    for p in progress_data:
        p.total_students = total_students
        exam_reports = StudentReport.objects.filter(exam=p.exam).order_by('-average_marks')
        for pos, r in enumerate(exam_reports, 1):
            if r.id == p.id:
                p.position = pos
                break
    breakdown = ExamResult.objects.filter(student=student, exam=report.exam).select_related('subject').order_by('subject__name')
    context = {
        'report': report,
        'student': student,
        'progress_data': progress_data,
        'breakdown': breakdown,
        'academic_year': academic_year,
        'class_room': class_room
    }
    return render(request, "academics/report_detail.html", context)

def report_list(request, pk):
    class_room = get_object_or_404(ClassRoom, pk=pk)
    academic_year = AcademicYear.objects.filter(is_active=True).first()
    reports = StudentReport.objects.filter(student__class_room=class_room, academic_year=academic_year).select_related('student', 'exam').order_by('student__last_name', 'exam__term__id')
    context = {'class_room': class_room, 'reports': reports, 'academic_year': academic_year}
    return render(request, "academics/report_list.html", context)

def generate_reports(request, class_id):
    class_room = get_object_or_404(ClassRoom, pk=class_id)
    active_exam = Exam.objects.filter(class_room=class_room, is_active=True).first()
    if not active_exam:
        messages.error(request, "No active exam found")
        return redirect('academics:class_dashboard', pk=class_id)
    students = Student.objects.filter(class_room=class_room, status="ACTIVE", is_active=True)
    for student in students:
        results = ExamResult.objects.filter(student=student, exam=active_exam)
        all_points = []
        points_for_best = []
        for result in results:
            avg = float(result.average_marks)
            if class_room.level_group == 'ORDINARY':
                p = get_ordinary_points(avg)
                result.points = p
                all_points.append(p)
                points_for_best.append(p)
            elif class_room.level_group == 'ADVANCED':
                p = get_advanced_points(avg)
                result.points = p
                all_points.append(p)
                if result.subject.name.upper()!= 'GENERAL STUDIES':
                    points_for_best.append(p)
            result.save()
        best_sum, division = calculate_best_points(points_for_best, class_room.level_group)
        total = results.aggregate(Sum('total_marks'))['total_marks__sum'] or 0
        avg = results.aggregate(Avg('average_marks'))['average_marks__avg'] or 0
        total_all = sum(all_points)
        StudentReport.objects.filter(student=student, exam=active_exam).delete()
        StudentReport.objects.create(
            student=student, exam=active_exam, academic_year=active_exam.term.academic_year,
            total_marks=total, average_marks=avg,
            total_points=total_all,
            division_points=best_sum,
            division=division
        )
    reports = StudentReport.objects.filter(exam=active_exam).order_by('-average_marks')
    for position, report in enumerate(reports, start=1):
        report.position = position
        report.save()
    messages.success(request, f"Reports regenerated for {active_exam.name}")
    return redirect('academics:class_dashboard', pk=class_id)

def bulk_print_class_reports(request, class_id, exam_id):
    class_room = get_object_or_404(ClassRoom, id=class_id)
    exam = get_object_or_404(Exam, id=exam_id)
    students = Student.objects.filter(class_room=class_room, status="ACTIVE", is_active=True).order_by('admission_number')
    result_data = []
    level_group = class_room.level_group
    for student in students:
        report = StudentReport.objects.filter(student=student, exam=exam).first()
        subject_results = ExamResult.objects.filter(student=student, exam=exam).select_related('subject').order_by('subject__name')
        result_data.append({'student': student, 'report': report, 'subjects': subject_results})
    context = {
        'school_name': "YOUR SCHOOL NAME",
        'class_room': class_room,
        'exam': exam,
        'result_data': result_data,
        'level_group': level_group,
    }
    return render(request, 'academics/result_sheet.html', context)

def report_pdf(request, pk):
    messages.info(request, "PDF coming soon. Use Print button for now.")
    return redirect('academics:report_detail', pk=pk)

# ========== WHATSAPP VIEWS ==========
def send_results_to_parents(request, class_id, term_number):
    class_room = get_object_or_404(ClassRoom, id=class_id)
    academic_year = AcademicYear.objects.filter(is_active=True).first()
    if not academic_year:
        messages.error(request, "Hakuna Mwaka wa Masomo uliowekwa kuwa Active")
        return redirect('academics:class_dashboard', pk=class_id)
    term_map = {1: ["TERM1", "TERM 1"], 2: ["TERM2", "TERM 2"], 3: ["TERM3", "TERM 3"], 4: ["TERM4", "TERM 4"]}
    term_names = term_map.get(term_number)
    term = Term.objects.filter(academic_year=academic_year, name__in=term_names).first()
    if not term:
        messages.error(request, f"Term {term_names} haipatikani. Nenda Admin uunde Term ya '{term_names[0]}'")
        return redirect('academics:class_dashboard', pk=class_id)
    exams = Exam.objects.filter(term=term, class_room=class_room, is_active=True).order_by('id')
    if not exams.exists():
        messages.error(request, f"Hakuna Exam iliyowekwa Active kwa {term.name} kwenye {class_room}")
        return redirect('academics:class_dashboard', pk=class_id)
    exam = exams.first()
    reports = list(StudentReport.objects.filter(exam=exam, student__class_room=class_room).select_related('student', 'student__guardian').order_by('-average_marks'))
    if not reports:
        messages.error(request, f"Hakuna ripoti kwa {exam.name}. Kwanza fanya 'Bulk Upload' au 'Generate Reports'")
        return redirect('academics:class_dashboard', pk=class_id)
    total_students = len(reports)
    for pos, report in enumerate(reports, start=1):
        report.position = pos
    term_label = {1: 'Midterm 1', 2: 'Terminal', 3: 'Midterm 2', 4: 'Annual'}.get(term_number)
    message_data = []
    domain = "https://the-system-otxf.onrender.com" # FIXED: Hakuna /64/ tena

    for report in reports:
        guardian = report.student.guardian
        if not guardian or not guardian.phone: continue
        phone = str(guardian.phone).replace(" ", "").replace("-", "").replace("+", "")
        if phone.startswith('0'): phone = '255' + phone[1:]
        elif not phone.startswith('255'): phone = '255' + phone
        student_name = _get_student_name(report.student)
        result_url = f"{domain}{reverse('academics:report_detail', args=[report.id])}" # FIXED: Link sahihi

        message = f"""*BABY A SCHOOL*
Ndugu mzazi wa {student_name}

MATOKEO YA {term.name} {term_label}
Nafasi: {report.position}/{total_students}
Wastani: {report.average_marks:.2f}%

*Tazama Ripoti Kamili:*
{result_url}

Asante."""
        wa_link = f"https://wa.me/{phone}?text={quote(message)}"
        message_data.append({'name': student_name, 'phone': phone, 'message': message, 'link': wa_link})
    context = {
        'class_room': class_room,
        'message_data': message_data,
        'term_info': f"{term.name} {term_label}",
        'total': len(message_data)
    }
    return render(request, 'academics/send_whatsapp.html', context)

def send_whatsapp_page(request, class_id):
    class_room = get_object_or_404(ClassRoom, id=class_id)
    links = request.session.get('whatsapp_links', [])
    term_info = request.session.get('term_info', 'Results')
    if not links:
        messages.warning(request, "No links to send. Check errors above.")
    context = {'class_room': class_room, 'links': links, 'term_info': term_info, 'total': len(links)}
    return render(request, 'academics/send_whatsapp.html', context)

# ========== MARKS ENTRY ==========
@login_required
def enter_marks(request, class_pk, exam_pk):
    class_room = get_object_or_404(ClassRoom, id=class_pk)
    exam = get_object_or_404(Exam, id=exam_pk)
    students = Student.objects.filter(class_room=class_room).order_by('first_name')
    subjects = Subject.objects.filter(class_room=class_room)
    if request.method == 'POST':
        for student in students:
            for subject in subjects:
                test_marks = request.POST.get(f'test_{student.id}_{subject.id}', 0)
                exam_marks = request.POST.get(f'exam_{student.id}_{subject.id}', 0)
                ExamResult.objects.update_or_create(
                    student=student,
                    exam=exam,
                    subject=subject,
                    defaults={'test_marks': test_marks, 'exam_marks': exam_marks}
                )
        messages.success(request, 'Alama zimehifadhiwa kikamilifu!')
        return redirect('academics:class_detail', pk=class_pk)
    context = {'class_room': class_room, 'exam': exam, 'students': students, 'subjects': subjects}
    return render(request, 'academics/enter_marks.html', context)